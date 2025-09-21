from datetime import datetime
import tempfile
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from app.niceGUI_folder.header import get_header
from app.niceGUI_folder.auth_middleware import require_auth
from app.niceGUI_folder.auth_service import AuthService
from nicegui import ui
from app.database_folder.orm import AsyncOrm


studbook_table_columns = [
    {'name': 'lfd_nr', 'label': 'Lfd Nr', 'field': 'lfd_nr', 'align': 'center', 'sortable': True},
    {'name': 'datum', 'label': 'Datum', 'field': 'datum', 'align': 'center', 'sortable': True},
    {'name': 'tiername', 'label': 'Tiername', 'field': 'tiername', 'align': 'left', 'sortable': True},
    {'name': 'zb_nummer', 'label': 'ZB-Nummer', 'field': 'zb_nummer', 'align': 'center', 'sortable': True},
    {'name': 'microchip', 'label': 'Microchip', 'field': 'microchip', 'align': 'center', 'sortable': True},
    {'name': 'geburtsdatum', 'label': 'Geburtsdatum', 'field': 'geburtsdatum', 'align': 'center', 'sortable': True},
    {'name': 'gender', 'label': 'w/m', 'field': 'gender', 'align': 'center', 'sortable': True},
    {'name': 'rasse', 'label': 'Rasse', 'field': 'rasse', 'align': 'left', 'sortable': True},
    {'name': 'ausgabe', 'label': 'Ausgabe', 'field': 'ausgabe', 'align': 'left', 'sortable': True},
    {'name': 'besitzer', 'label': 'Besitzer', 'field': 'besitzer', 'align': 'left', 'sortable': True},
    {'name': 'kommentar', 'label': 'Kommentar', 'field': 'kommentar', 'align': 'left', 'sortable': True},
    {'name': 'wcf_sticker', 'label': 'WCF Sticker', 'field': 'wcf_sticker', 'align': 'center', 'sortable': True},
]


@require_auth(required_permission=2)
async def studbook_page_render(current_user=None, session_id=None):
    """Render the Studbook page with grouped structure by breeder and litter"""

    _, cats_data = await AsyncOrm.get_cat_info()
    _, owners_data = await AsyncOrm.get_owner()
    _, breeds_data = await AsyncOrm.get_breed()

    owner_filter = AuthService.get_user_cats_filter(current_user)
    if owner_filter is not None:
        if owner_filter == -1:
            cats_data = []
        else:
            cats_data = [cat for cat in cats_data if cat.get('owner_id') == owner_filter]
    breeder_options = {
        breed['breed_id']: f"{breed['breed_firstname']} {breed['breed_surname']}"
        for breed in breeds_data
        if breed['breed_firstname'] and breed['breed_surname']
    }
    owner_options = {
        owner['owner_id']: f"{owner['owner_firstname']} {owner['owner_surname']}"
        for owner in owners_data
        if owner['owner_firstname'] and owner['owner_surname']
    }
    ems_color_options = list(set([cat.get('colour') for cat in cats_data if cat.get('colour')]))
    status_options = list(set([cat.get('status') for cat in cats_data if cat.get('status')]))

    # Filter variables
    filter_inputs = {}
    results_label = None
    studbook_container = None

    def apply_filters():
        """Apply all filters to the data"""
        if not cats_data:
            return []

        filtered_cats = cats_data.copy()

        # Search filter
        if filter_inputs.get('search_input') and filter_inputs['search_input'].value:
            search_term = filter_inputs['search_input'].value.lower()
            filtered_cats = [
                cat for cat in filtered_cats
                if (search_term in str(cat.get('firstname', '') or '').lower() or
                    search_term in str(cat.get('surname', '') or '').lower() or
                    search_term in str(cat.get('callname', '') or '').lower() or
                    search_term in str(cat.get('microchip', '') or '').lower() or
                    search_term in str(cat.get('haritage_number', '') or '').lower() or
                    search_term in str(cat.get('haritage_number_2', '') or '').lower())
            ]

        # Breeder filter
        if filter_inputs.get('breeder_filter') and filter_inputs['breeder_filter'].value:
            selected_breeder_name = filter_inputs['breeder_filter'].value
            selected_breeder_id = next((k for k, v in breeder_options.items() if v == selected_breeder_name), None)
            if selected_breeder_id:
                filtered_cats = [cat for cat in filtered_cats if cat.get('breed') == selected_breeder_id]

        # Owner filter
        if filter_inputs.get('owner_filter_select') and filter_inputs['owner_filter_select'].value:
            selected_owner_name = filter_inputs['owner_filter_select'].value
            selected_owner_id = next((k for k, v in owner_options.items() if v == selected_owner_name), None)
            if selected_owner_id:
                filtered_cats = [cat for cat in filtered_cats if cat.get('owner_id') == selected_owner_id]

        # EMS color filter
        if filter_inputs.get('ems_color_filter') and filter_inputs['ems_color_filter'].value:
            filtered_cats = [cat for cat in filtered_cats
                             if cat.get('colour') == filter_inputs['ems_color_filter'].value]

        # Status filter
        if filter_inputs.get('status_filter') and filter_inputs['status_filter'].value:
            filtered_cats = [cat for cat in filtered_cats if cat.get('status') == filter_inputs['status_filter'].value]

        # Birthday range filter
        if filter_inputs.get('birthday_from') and filter_inputs['birthday_from'].value:
            try:
                from_date = datetime.strptime(filter_inputs['birthday_from'].value, '%Y-%m-%d').date()
                filtered_cats = [cat for cat in filtered_cats
                                 if cat.get('birthday') and cat.get('birthday') >= from_date]
                print(f"DEBUG: Filtered by birthday_from: {from_date}, found {len(filtered_cats)} cats")
            except ValueError as e:
                print(f"DEBUG: Error parsing birthday_from: {e}")

        if filter_inputs.get('birthday_to') and filter_inputs['birthday_to'].value:
            try:
                to_date = datetime.strptime(filter_inputs['birthday_to'].value, '%Y-%m-%d').date()
                filtered_cats = [cat for cat in filtered_cats
                                 if cat.get('birthday') and cat.get('birthday') <= to_date]
                print(f"DEBUG: Filtered by birthday_to: {to_date}, found {len(filtered_cats)} cats")
            except ValueError as e:
                print(f"DEBUG: Error parsing birthday_to: {e}")

        return filtered_cats

    def group_cats_by_breeder_and_litter(filtered_cats):
        """Group cats by breeder and then by litter"""
        grouped = {}

        for cat in filtered_cats:
            breeder_name = f"{cat.get('breed_firstname', '')} {cat.get('breed_surname', '')}".strip()
            if not breeder_name:
                breeder_name = "Unknown Breeder"

            if breeder_name not in grouped:
                grouped[breeder_name] = {
                    "Zuchttiere": [],
                    "litters": {}
                }

            # Check if it's a breeding animal
            if cat.get('breeding_animal'):
                grouped[breeder_name]["Zuchttiere"].append(cat)
            else:
                # Group by litter
                litter = cat.get('litter', 'No Litter')
                if litter not in grouped[breeder_name]["litters"]:
                    grouped[breeder_name]["litters"][litter] = []
                grouped[breeder_name]["litters"][litter].append(cat)

        return grouped

    def create_studbook_row(cat, lfd_nr):
        """Create a row for the studbook table"""
        # Format cat name
        tiername = f"{cat.get('firstname', '')} {cat.get('surname', '')}".strip()

        # Format registration numbers
        reg_numbers = []
        if cat.get('haritage_number'):
            reg_numbers.append(cat.get('haritage_number'))
        if cat.get('haritage_number_2'):
            reg_numbers.append(cat.get('haritage_number_2'))
        zb_nummer = ', '.join(reg_numbers) if reg_numbers else ''

        # Format birthday
        geburtsdatum = cat.get('birthday').strftime('%Y-%m-%d') if cat.get('birthday') else ''

        # Format gender (w/m)
        gender = 'w' if cat.get('gender') == 'Female' else 'm' if cat.get('gender') == 'Male' else ''

        # Format breed name
        rasse = f"{cat.get('breed_firstname', '')} {cat.get('breed_surname', '')}".strip()

        # Determine document type
        ausgabe = 'Stammbaum' if cat.get('breeding_animal') else 'Abschrift'

        # Format owner name
        besitzer = f"{cat.get('owner_firstname', '')} {cat.get('owner_surname', '')}".strip()

        # Get comments
        kommentar = cat.get('notes', '') or ''

        # WCF Sticker (placeholder for future implementation)
        wcf_sticker = '✓' if cat.get('wcf_sticker') else ''

        return {
            'lfd_nr': lfd_nr,
            'datum': datetime.now().strftime('%Y-%m-%d'),
            'tiername': tiername or 'Not specified',
            'zb_nummer': zb_nummer or 'Not specified',
            'microchip': cat.get('microchip', '') or 'Not specified',
            'geburtsdatum': geburtsdatum or 'Not specified',
            'gender': gender,
            'rasse': rasse or 'Not specified',
            'ausgabe': ausgabe,
            'besitzer': besitzer or 'Not specified',
            'kommentar': kommentar,
            'wcf_sticker': wcf_sticker,
            'raw_data': cat  # Store full data for detail view
        }

    async def show_cat_details(cat_data):
        """Show detailed cat information in modal"""
        if not cat_data or 'raw_data' not in cat_data:
            return

        cat = cat_data['raw_data']

        with ui.dialog() as dialog, ui.card().classes('w-full max-w-4xl'):
            ui.markdown(f"## 📋 Studbook - {cat.get('firstname', '')} {cat.get('surname', '')}")

            with ui.grid(columns=2):
                # Basic Information
                with ui.card():
                    ui.markdown("### 🐱 Basic Information")
                    ui.label(f"**Name:** {cat.get('firstname', 'Not specified')} {cat.get('surname', '')}")
                    ui.label(f"**Callname:** {cat.get('callname', 'Not specified')}")
                    ui.label(f"**Gender:** {cat.get('gender', 'Not specified')}")
                    ui.label(f"**Birthday:** {cat.get('birthday', 'Not specified')}")
                    ui.label(f"**Microchip:** {cat.get('microchip', 'Not specified')}")
                    ui.label(f"**EMS Color:** {cat.get('colour', 'Not specified')}")
                    ui.label(f"**Litter:** {cat.get('litter', 'Not specified')}")

                # Registration Information
                with ui.card():
                    ui.markdown("### 📜 Registration Information")
                    ui.label(f"**Registration Number 1:** {cat.get('haritage_number', 'Not specified')}")
                    ui.label(f"**Registration Number 2:** {cat.get('haritage_number_2', 'Not specified')}")
                    ui.label(f"**Breeding Animal:** {'Yes' if cat.get('breeding_animal') else 'No'}")
                    ui.label(f"**Status:** {cat.get('status', 'Not specified')}")
                    ui.label(f"**Document Type:** {'Stammbaum' if cat.get('breeding_animal') else 'Abschrift'}")
                    ui.label(f"**WCF Sticker:** {'✓' if cat.get('wcf_sticker') else 'No'}")

                # Breeder Information
                with ui.card():
                    ui.markdown("### 👨‍🌾 Breeder Information")
                    ui.label(f"**Breeder:** {cat.get('breed_firstname', '')} {cat.get('breed_surname', '')}")
                    ui.label(f"**Email:** {cat.get('breed_email', 'Not specified')}")
                    ui.label(f"**Phone:** {cat.get('breed_phone', 'Not specified')}")
                    ui.label(f"**Country:** {cat.get('breed_country', 'Not specified')}")
                    ui.label(f"**City:** {cat.get('breed_city', 'Not specified')}")

                # Owner Information
                with ui.card():
                    ui.markdown("### 👤 Owner Information")
                    ui.label(f"**Owner:** {cat.get('owner_firstname', '')} {cat.get('owner_surname', '')}")
                    ui.label(f"**Email:** {cat.get('owner_email', 'Not specified')}")
                    ui.label(f"**Country:** {cat.get('owner_country', 'Not specified')}")
                    ui.label(f"**City:** {cat.get('owner_city', 'Not specified')}")

            # Additional Information
            if cat.get('notes'):
                with ui.card():
                    ui.markdown("### 📝 Comments")
                    ui.label(f"**Notes:** {cat.get('notes', '')}")

            if cat.get('dam') or cat.get('sire'):
                with ui.card():
                    ui.markdown("### 👨‍👩‍👧‍👦 Parents")
                    ui.label(f"**Dam:** {cat.get('dam', 'Not specified')}")
                    ui.label(f"**Sire:** {cat.get('sire', 'Not specified')}")

            ui.button('Close', on_click=dialog.close).props('color=primary')

        dialog.open()

    async def render_studbook_structure():
        """Render the grouped studbook structure"""
        filtered_cats = apply_filters()
        grouped_data = group_cats_by_breeder_and_litter(filtered_cats)

        lfd_nr_counter = 1

        for breeder_name, breeder_data in grouped_data.items():
            # Breeder expansion
            with ui.expansion(breeder_name, icon='person').classes('w-full'):

                # Zuchttiere section
                if breeder_data["Zuchttiere"]:
                    with ui.expansion("🐾 Zuchttiere", icon='pets').classes('w-full'):
                        table = ui.table(
                            columns=studbook_table_columns,
                            rows=[create_studbook_row(cat, lfd_nr_counter + i)
                                  for i, cat in enumerate(breeder_data["Zuchttiere"])]
                        ).classes('w-full')
                        lfd_nr_counter += len(breeder_data["Zuchttiere"])

                        # Set up table row click handler
                        table.on('rowClick', lambda e: show_cat_details(e.args[1]))

                # Litters sections
                for litter_name, litter_cats in breeder_data["litters"].items():
                    with ui.expansion(f"👶 {litter_name}", icon='child_care').classes('w-full'):
                        table = ui.table(
                            columns=studbook_table_columns,
                            rows=[create_studbook_row(cat, lfd_nr_counter + i) for i, cat in enumerate(litter_cats)]
                        ).classes('w-full')
                        lfd_nr_counter += len(litter_cats)

                        # Set up table row click handler
                        table.on('rowClick', lambda e: show_cat_details(e.args[1]))

    async def export_to_xlsx():
        """Export filtered studbook data to XLSX file"""
        try:
            ui.notify('Export started...', type='info', position='top')
            
            # Get current filtered data
            filtered_cats = apply_filters()
            
            if not filtered_cats:
                ui.notify('No data to export', type='warning', position='top')
                return
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Studbook Export"
            
            # Define headers based on studbook_table_columns
            headers = [col['label'] for col in studbook_table_columns]
            
            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Write data rows
            for row_idx, cat in enumerate(filtered_cats, 2):
                # Create studbook row data
                studbook_row = create_studbook_row(cat, row_idx - 1)
                
                # Write each column value
                for col_idx, col_def in enumerate(studbook_table_columns, 1):
                    field_name = col_def['field']
                    value = studbook_row.get(field_name, '')
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Generate filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'studbook_export_{timestamp}.xlsx'
            
            # Save to temporary file and read content
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                wb.save(tmp_file.name)
                temp_path = tmp_file.name
            
            # Read file content
            with open(temp_path, 'rb') as f:
                file_content = f.read()
            
            # Clean up temporary file
            os.unlink(temp_path)
            
            # Download the file using ui.download with content
            ui.download(file_content, filename)
            
            ui.notify('XLSX file ready for download', type='positive', position='top')
            
        except Exception as e:
            print(f"Export error: {e}")
            ui.notify(f'Export failed: {str(e)}', type='negative', position='top')

    async def export_to_pdf():
        """Export filtered studbook data to PDF file"""
        try:
            ui.notify('PDF export started...', type='info', position='top')
            
            # Get current filtered data
            filtered_cats = apply_filters()
            
            if not filtered_cats:
                ui.notify('No data to export', type='warning', position='top')
                return
            
            # Generate filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'studbook_export_{timestamp}.pdf'
            
            # Create temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                temp_path = tmp_file.name
            
            # Create PDF document
            doc = SimpleDocTemplate(temp_path, pagesize=landscape(A4))
            story = []
            
            # Define styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            
            # Add title
            title = Paragraph("Studbook Export", title_style)
            story.append(title)
            story.append(Spacer(1, 12))
            
            # Prepare table data
            headers = [col['label'] for col in studbook_table_columns]
            table_data = [headers]
            
            # Add data rows
            for cat in filtered_cats:
                studbook_row = create_studbook_row(cat, len(table_data))
                row_data = []
                for col_def in studbook_table_columns:
                    field_name = col_def['field']
                    value = str(studbook_row.get(field_name, ''))
                    # Truncate long text for PDF display
                    if len(value) > 20:
                        value = value[:17] + "..."
                    row_data.append(value)
                table_data.append(row_data)
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                # Header row styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Data rows styling
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            story.append(table)
            
            # Build PDF
            doc.build(story)
            
            # Read file content
            with open(temp_path, 'rb') as f:
                file_content = f.read()
            
            # Clean up temporary file
            os.unlink(temp_path)
            
            # Download the file
            ui.download(file_content, filename)
            
            ui.notify('PDF file ready for download', type='positive', position='top')
            
        except Exception as e:
            print(f"PDF export error: {e}")
            ui.notify(f'PDF export failed: {str(e)}', type='negative', position='top')

    async def clear_all_filters():
        """Clear all filter inputs"""
        for key, input_widget in filter_inputs.items():
            if input_widget:
                input_widget.value = ''
        await update_studbook_display()

    async def render_page_content():
        """Render the main page content"""
        # Render page
        get_header("Studbook")

        ui.markdown("## 📚 Studbook")
        ui.markdown("Official registry of registered cats and their litters")

        # Filters section - compact design like cats_page
        with ui.card().classes('w-full q-pa-md q-mb-md'):
            ui.label('🔍 Filters').classes('text-h6 q-mb-md')

            with ui.grid(columns=4).classes('gap-4 w-full'):
                # Search
                filter_inputs['search_input'] = ui.input(
                    label='Search (Name, microchip, ZB...)'
                ).props('outlined dense')

                # Breeder filter
                filter_inputs['breeder_filter'] = ui.select(
                    options=[''] + list(breeder_options.values()),
                    label='Breeder'
                ).props('outlined dense')

                # Owner filter
                filter_inputs['owner_filter_select'] = ui.select(
                    options=[''] + list(owner_options.values()),
                    label='Owner'
                ).props('outlined dense')

                # EMS color filter
                filter_inputs['ems_color_filter'] = ui.select(
                    options=[''] + ems_color_options,
                    label='EMS Color'
                ).props('outlined dense')

                # Status filter
                filter_inputs['status_filter'] = ui.select(
                    options=[''] + status_options,
                    label='Status'
                ).props('outlined dense')

                # Birthday filters
                filter_inputs['birthday_from'] = ui.input(label='Birthday From').props('type=date outlined dense')
                filter_inputs['birthday_to'] = ui.input(label='Birthday To').props('type=date outlined dense')

                # Clear filters button
                clear_filters_btn = ui.button('Clear All', icon='clear').props('color=secondary outline')
                
                # Export buttons
                export_xlsx_btn = ui.button('Export XLSX', icon='download').props('color=primary')
                export_pdf_btn = ui.button('Export PDF', icon='picture_as_pdf').props('color=secondary')

        # Set up event handlers for the newly created filters
        filter_inputs['search_input'].on_value_change(lambda: update_studbook_display())
        filter_inputs['breeder_filter'].on_value_change(lambda: update_studbook_display())
        filter_inputs['owner_filter_select'].on_value_change(lambda: update_studbook_display())
        filter_inputs['ems_color_filter'].on_value_change(lambda: update_studbook_display())
        filter_inputs['status_filter'].on_value_change(lambda: update_studbook_display())
        filter_inputs['birthday_from'].on_value_change(lambda: update_studbook_display())
        filter_inputs['birthday_to'].on_value_change(lambda: update_studbook_display())
        clear_filters_btn.on_click(clear_all_filters)
        export_xlsx_btn.on_click(export_to_xlsx)
        export_pdf_btn.on_click(export_to_pdf)

        # Results counter
        nonlocal results_label
        results_label = ui.label().classes('q-mb-md')

        # Studbook container
        nonlocal studbook_container
        studbook_container = ui.column().classes('w-full')

        # Initial update
        await update_studbook_display()

    async def update_studbook_display():
        """Update the studbook display without full page reload"""
        # Update only the results section, not the entire page
        filtered_cats = apply_filters()

        # Debug: print filter values
        print("DEBUG: Filter values:")
        for key, widget in filter_inputs.items():
            if widget:
                print(f"  {key}: '{widget.value}'")

        # Update results counter
        if results_label:
            results_label.text = f'Records found: {len(filtered_cats)}'

        # Clear and rebuild only the studbook structure
        if studbook_container:
            studbook_container.clear()
            with studbook_container:
                await render_studbook_structure()

    # Initial render
    await render_page_content()
