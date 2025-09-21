from datetime import date, datetime
import tempfile
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from app.niceGUI_folder.header import get_header
from app.niceGUI_folder.auth_middleware import require_auth
from app.niceGUI_folder.auth_service import AuthService
from nicegui import ui
from app.database_folder.orm import AsyncOrm

cats_column = [
    {'name': 'id', 'label': 'ID', 'field': 'id', 'align': 'left'},
    {'name': 'firstname', 'label': 'Name', 'field': 'firstname', 'align': 'left'},
    {'name': 'surname', 'label': 'Surname', 'field': 'surname', 'align': 'left'},
    {'name': 'callname', 'label': 'Callname', 'field': 'callname', 'align': 'left'},
    {'name': 'gender', 'label': 'Gender', 'field': 'gender', 'align': 'left'},
    {'name': 'birthday', 'label': 'Birthday', 'field': 'birthday', 'align': 'left'},
    {'name': 'microchip', 'label': 'Microchip', 'field': 'microchip', 'align': 'left'},
    {'name': 'title', 'label': 'Title', 'field': 'title', 'align': 'left'},
    {'name': 'haritage_number', 'label': 'Studbook 1', 'field': 'haritage_number', 'align': 'left'},
    {'name': 'haritage_number_2', 'label': 'Studbook 2', 'field': 'haritage_number_2', 'align': 'left'},
    {'name': 'eye_colour', 'label': 'Eye Color', 'field': 'eye_colour', 'align': 'left'},
    {'name': 'hair_type', 'label': 'Hair Type', 'field': 'hair_type', 'align': 'left'},
    {'name': 'colour', 'label': 'Color', 'field': 'colour', 'align': 'left'},
    {'name': 'litter', 'label': 'Litter', 'field': 'litter', 'align': 'left'},
    {'name': 'dam', 'label': 'Dam', 'field': 'dam', 'align': 'left'},
    {'name': 'sire', 'label': 'Sire', 'field': 'sire', 'align': 'left'},
    {'name': 'litter_size_male', 'label': 'Litter Male', 'field': 'litter_size_male', 'align': 'left'},
    {'name': 'litter_size_female', 'label': 'Litter Female', 'field': 'litter_size_female', 'align': 'left'},
    {'name': 'tests', 'label': 'Tests', 'field': 'tests', 'align': 'left'},
    {'name': 'blood_group', 'label': 'Blood Group', 'field': 'blood_group', 'align': 'left'},
    {'name': 'gencode', 'label': 'Gencode', 'field': 'gencode', 'align': 'left'},
    {'name': 'weight', 'label': 'Weight', 'field': 'weight', 'align': 'left'},
    {'name': 'birth_weight', 'label': 'Birth Weight', 'field': 'birth_weight', 'align': 'left'},
    {'name': 'transfer_weight', 'label': 'Transfer Weight', 'field': 'transfer_weight', 'align': 'left'},
    {'name': 'breeding_lock', 'label': 'Breeding Lock', 'field': 'breeding_lock', 'align': 'left'},
    {'name': 'breeding_animal', 'label': 'Breeding Animal', 'field': 'breeding_animal', 'align': 'left'},
    {'name': 'kitten_transfer', 'label': 'Kitten Transfer', 'field': 'kitten_transfer', 'align': 'left'},
    {'name': 'wcf_sticker', 'label': 'WCF Sticker', 'field': 'wcf_sticker', 'align': 'left'},
    {'name': 'birth_country', 'label': 'Birth Country', 'field': 'birth_country', 'align': 'left'},
    {'name': 'location', 'label': 'Location', 'field': 'location', 'align': 'left'},
    {'name': 'association', 'label': 'Association', 'field': 'association', 'align': 'left'},
    {'name': 'faults_deviations', 'label': 'Faults', 'field': 'faults_deviations', 'align': 'left'},
    {'name': 'jaw_fault', 'label': 'Jaw Fault', 'field': 'jaw_fault', 'align': 'left'},
    {'name': 'hernia', 'label': 'Hernia', 'field': 'hernia', 'align': 'left'},
    {'name': 'testicles', 'label': 'Testicles', 'field': 'testicles', 'align': 'left'},
    {'name': 'death_date', 'label': 'Death Date', 'field': 'death_date', 'align': 'left'},
    {'name': 'death_cause', 'label': 'Death Cause', 'field': 'death_cause', 'align': 'left'},
    {'name': 'status', 'label': 'Status', 'field': 'status', 'align': 'left'},
    {'name': 'features', 'label': 'Features', 'field': 'features', 'align': 'left'},
    {'name': 'notes', 'label': 'Notes', 'field': 'notes', 'align': 'left'},
    {'name': 'show_results', 'label': 'Show Results', 'field': 'show_results', 'align': 'left'},
    {'name': 'description', 'label': 'Description', 'field': 'description', 'align': 'left'},
    {'name': 'breed_name', 'label': 'Breeder', 'field': 'breed_name', 'align': 'left'},
    {'name': 'owner_name', 'label': 'Owner', 'field': 'owner_name', 'align': 'left'},
    {'name': 'actions', 'label': 'Actions', 'field': 'actions', 'align': 'center'},
]


def get_edit_button_vue(current_user):
    can_edit = current_user and current_user.get('owner_permission') == 1

    if can_edit:
        return r'''
          <q-tr :props="props">
            <q-td v-for="col in props.cols" :key="col.name" :props="props">
              <template v-if="col.name === 'actions'">
                <q-btn size="sm" color="primary" flat
                       :href="'/cat_profile/' + props.row.id"
                       label="View" class="q-mr-xs" />
                <q-btn size="sm" color="secondary" flat
                       :href="'/edit_cat/' + props.row.id"
                       label="Edit" />
              </template>
              <template v-else-if="col.name === 'firstname'">
                <q-btn flat no-caps color="primary"
                       :href="'/cat_profile/' + props.row.id"
                       :label="col.value" />
              </template>
              <template v-else>
                {{ col.value }}
              </template>
            </q-td>
          </q-tr>
        '''
    else:
        return r'''
          <q-tr :props="props">
            <q-td v-for="col in props.cols" :key="col.name" :props="props">
              <template v-if="col.name === 'actions'">
                <q-btn size="sm" color="primary" flat
                       :href="'/cat_profile/' + props.row.id"
                       label="View" />
              </template>
              <template v-else-if="col.name === 'firstname'">
                <q-btn flat no-caps color="primary"
                       :href="'/cat_profile/' + props.row.id"
                       :label="col.value" />
              </template>
              <template v-else>
                {{ col.value }}
              </template>
            </q-td>
          </q-tr>
        '''


async def get_cats_rows(cats_rows):
    safe_rows = []
    for r in cats_rows:
        row = dict(r)
        b = row.get('birthday')
        if isinstance(b, (date, datetime)):
            row['birthday'] = b.isoformat()
        row.setdefault('actions', '')
        safe_rows.append(row)
    return safe_rows


@require_auth(required_permission=2)
async def cats_page_render(current_user=None, session_id=None):
    get_header('🐱 Cats')

    # Add Cat button for admins
    if current_user and current_user.get('owner_permission') == 1:
        with ui.row().classes('q-pa-md'):
            ui.button('Add Cat', on_click=lambda: ui.navigate.to('/add_cat')).classes('q-mr-sm')

    # Load data for filters (only metadata, not all data)
    _, owners_data = await AsyncOrm.get_owner()
    _, breeds_data = await AsyncOrm.get_breed()

    # Pagination settings
    PAGE_SIZE = 100
    current_offset = 0
    all_cats_data = []
    loading_more = False
    has_more_data = True

    # Apply user permission filter
    owner_filter = AuthService.get_user_cats_filter(current_user)
    if owner_filter is not None and owner_filter == -1:
        cats_data = []
    else:
        cats_data = []

    async def load_cats_data(offset=0, limit=PAGE_SIZE, reset=False):
        """Load cats data with pagination"""
        nonlocal all_cats_data, has_more_data, loading_more
        
        if loading_more:
            return []
            
        loading_more = True
        
        try:
            # Load data from database with pagination
            _, new_cats = await AsyncOrm.get_cat_info(limit=limit, offset=offset)
            
            # Apply user permission filter
            if owner_filter is not None and owner_filter != -1:
                new_cats = [cat for cat in new_cats if cat.get('owner_id') == owner_filter]
            
            if reset:
                all_cats_data = new_cats
            else:
                all_cats_data.extend(new_cats)
            
            # Check if there's more data
            has_more_data = len(new_cats) == limit
            
            return new_cats
        finally:
            loading_more = False

    # Create filter options (will be updated after first load)
    gender_options = ['Male', 'Female']
    eye_color_options = []
    hair_type_options = []
    status_options = list(set([cat.get('status') for cat in cats_data if cat.get('status')]))
    
    owner_options = {owner['owner_id']: f"{owner['owner_firstname']} {owner['owner_surname']}"
                     for owner in owners_data}
    breed_options = {breed['breed_id']: f"{breed['breed_firstname']} {breed['breed_surname']}"
                     for breed in breeds_data}

    # Filter UI
    with ui.card().classes('w-full q-pa-md q-mb-md'):
        ui.label('🔍 Filters').classes('text-h6 q-mb-md')
        
        with ui.grid(columns=4).classes('gap-4 w-full'):
            # Search
            search_input = ui.input(label='Search (Name, Microchip, etc.)').props('outlined dense')
            
            # Gender filter
            gender_filter = ui.select(
                options=[''] + gender_options,
                label='Gender'
            ).props('outlined dense clearable')
            
            # Owner filter
            owner_filter_select = ui.select(
                options=[''] + list(owner_options.values()),
                label='Owner'
            ).props('outlined dense clearable')
            
            # Breeder filter
            breeder_filter = ui.select(
                options=[''] + list(breed_options.values()),
                label='Breeder'
            ).props('outlined dense clearable')
            
            # Eye color filter
            eye_color_filter = ui.select(
                options=[''] + eye_color_options,
                label='Eye Color'
            ).props('outlined dense clearable')
            
            # Hair type filter
            hair_type_filter = ui.select(
                options=[''] + hair_type_options,
                label='Hair Type'
            ).props('outlined dense clearable')
            
            # Status filter
            status_filter = ui.select(
                options=[''] + status_options,
                label='Status'
            ).props('outlined dense clearable')
            
            # Color filter
            color_options = list(set([cat.get('colour') for cat in cats_data if cat.get('colour')]))
            color_filter = ui.select(
                options=[''] + color_options,
                label='Color'
            ).props('outlined dense clearable')
            
            # Birthday range
            birthday_from = ui.input(label='Birthday From').props('type=date outlined dense')
            birthday_to = ui.input(label='Birthday To').props('type=date outlined dense')
            
            # Weight range
            weight_min = ui.number(label='Min Weight (kg)', value=None, min=0, max=50, step=0.1).props('outlined dense')
            weight_max = ui.number(label='Max Weight (kg)', value=None, min=0, max=50, step=0.1).props('outlined dense')
            
            # Breeding status
            breeding_animal_filter = ui.select(
                options=['', 'Yes', 'No'],
                label='Breeding Animal'
            ).props('outlined dense clearable')
            
            breeding_lock_filter = ui.select(
                options=['', 'Yes', 'No'],
                label='Breeding Lock'
            ).props('outlined dense clearable')
            
            # Clear filters button
            clear_filters_btn = ui.button('Clear All Filters', color='secondary').props('outline')
            
            # Export buttons
            export_xlsx_btn = ui.button('Export XLSX', icon='download').props('color=primary')
            export_pdf_btn = ui.button('Export PDF', icon='picture_as_pdf').props('color=secondary')

    # Results counter
    results_label = ui.label('').classes('text-subtitle2 q-mb-sm')

    # Table container
    table_container = ui.column()

    async def apply_filters(reset_pagination=True):
        """Apply all filters to the data with pagination support"""
        nonlocal current_offset, all_cats_data
        
        if reset_pagination:
            current_offset = 0
            # Load first batch with filters applied
            await load_cats_data(offset=0, limit=PAGE_SIZE, reset=True)
        
        filtered_cats = all_cats_data.copy()
        
        # Search filter
        search_term = search_input.value.lower() if search_input.value else ''
        if search_term:
            filtered_cats = [
                cat for cat in filtered_cats
                if (search_term in str(cat.get('firstname', '') or '').lower() or
                    search_term in str(cat.get('surname', '') or '').lower() or
                    search_term in str(cat.get('microchip', '') or '').lower() or
                    search_term in str(cat.get('callname', '') or '').lower() or
                    search_term in str(cat.get('haritage_number', '') or '').lower() or
                    search_term in str(cat.get('owner_firstname', '') or '').lower() or
                    search_term in str(cat.get('owner_surname', '') or '').lower() or
                    search_term in str(cat.get('breed_firstname', '') or '').lower() or
                    search_term in str(cat.get('breed_surname', '') or '').lower())
            ]
        
        # Gender filter
        if gender_filter.value:
            filtered_cats = [cat for cat in filtered_cats if cat.get('gender') == gender_filter.value]
        
        # Owner filter
        if owner_filter_select.value:
            selected_owner_name = owner_filter_select.value
            selected_owner_id = next((k for k, v in owner_options.items() if v == selected_owner_name), None)
            if selected_owner_id:
                filtered_cats = [cat for cat in filtered_cats if cat.get('owner_id') == selected_owner_id]
        
        # Breeder filter
        if breeder_filter.value:
            selected_breeder_name = breeder_filter.value
            selected_breeder_id = next((k for k, v in breed_options.items() if v == selected_breeder_name), None)
            if selected_breeder_id:
                filtered_cats = [cat for cat in filtered_cats if cat.get('breed') == selected_breeder_id]
        
        # Eye color filter
        if eye_color_filter.value:
            filtered_cats = [cat for cat in filtered_cats if cat.get('eye_colour') == eye_color_filter.value]
        
        # Hair type filter
        if hair_type_filter.value:
            filtered_cats = [cat for cat in filtered_cats if cat.get('hair_type') == hair_type_filter.value]
        
        # Status filter
        if status_filter.value:
            filtered_cats = [cat for cat in filtered_cats if cat.get('status') == status_filter.value]
        
        # Color filter
        if color_filter.value:
            filtered_cats = [cat for cat in filtered_cats if cat.get('colour') == color_filter.value]
        
        # Birthday range filter
        if birthday_from.value:
            try:
                from_date = datetime.strptime(birthday_from.value, '%Y-%m-%d').date()
                filtered_cats = [cat for cat in filtered_cats
                                 if cat.get('birthday') and cat.get('birthday') >= from_date]
            except ValueError:
                pass
        
        if birthday_to.value:
            try:
                to_date = datetime.strptime(birthday_to.value, '%Y-%m-%d').date()
                filtered_cats = [cat for cat in filtered_cats
                                 if cat.get('birthday') and cat.get('birthday') <= to_date]
            except ValueError:
                pass
        
        # Weight range filter
        if weight_min.value is not None:
            filtered_cats = [cat for cat in filtered_cats
                             if cat.get('weight') and cat.get('weight') >= weight_min.value]
        
        if weight_max.value is not None:
            filtered_cats = [cat for cat in filtered_cats
                             if cat.get('weight') and cat.get('weight') <= weight_max.value]
        
        # Breeding animal filter
        if breeding_animal_filter.value:
            breeding_value = breeding_animal_filter.value == 'Yes'
            filtered_cats = [cat for cat in filtered_cats if cat.get('breeding_animal') == breeding_value]
        
        # Breeding lock filter
        if breeding_lock_filter.value:
            lock_value = breeding_lock_filter.value == 'Yes'
            filtered_cats = [cat for cat in filtered_cats if cat.get('breeding_lock') == lock_value]
        
        return filtered_cats

    async def load_more_data():
        """Load more data when scrolling"""
        nonlocal current_offset, loading_more
        
        if loading_more or not has_more_data:
            return
            
        current_offset += PAGE_SIZE
        new_cats = await load_cats_data(offset=current_offset, limit=PAGE_SIZE, reset=False)
        
        if new_cats:
            # Update table with new data
            await update_table()

    async def update_table():
        """Update the table with filtered data"""
        filtered_cats = await apply_filters(reset_pagination=False)
        
        # Prepare rows for display
        display_rows = []
        for cat in filtered_cats:
            # Format title
            title_display = cat.get('title')[0] if cat.get('title') and len(cat.get('title')) > 0 else ''
            
            # Format owner name
            owner_name = f"{cat.get('owner_firstname', '')} {cat.get('owner_surname', '')}".strip()
            
            # Format breeder name
            breeder_name = f"{cat.get('breed_firstname', '')} {cat.get('breed_surname', '')}".strip()
            
            # Format birthday
            birthday_display = cat.get('birthday').isoformat() if cat.get('birthday') else ''
            
            row = {
                'id': cat.get('id'),
                'firstname': cat.get('firstname', ''),
                'surname': cat.get('surname', ''),
                'callname': cat.get('callname', ''),
                'gender': cat.get('gender', ''),
                'birthday': birthday_display,
                'microchip': cat.get('microchip', ''),
                'title': title_display,
                'eye_colour': cat.get('eye_colour', ''),
                'hair_type': cat.get('hair_type', ''),
                'colour': cat.get('colour', ''),
                'breed_name': breeder_name,
                'owner_name': owner_name,
                'status': cat.get('status', ''),
                'actions': ''
            }
            display_rows.append(row)
        
        # Update results counter
        results_label.text = f"Found {len(display_rows)} cats"
        
        # Update table
        table_container.clear()
        with table_container:
            if display_rows:
                table = ui.table(columns=cats_column, rows=display_rows, row_key='id').classes('q-pa-md')
                table.add_slot('body', get_edit_button_vue(current_user))
                
                # Add load more button if there's more data
                if has_more_data:
                    with ui.row().classes('q-pa-md justify-center'):
                        load_more_btn = ui.button('Load More', icon='expand_more',
                                                  on_click=load_more_data).props('color=primary outline')
                        if loading_more:
                            load_more_btn.props('loading')
                elif len(display_rows) > 0:
                    with ui.row().classes('q-pa-md justify-center'):
                        ui.label('All records loaded').classes('text-grey-6')
            else:
                ui.label('No cats found matching the criteria').classes('text-center q-py-xl text-grey-6')

    async def export_to_xlsx():
        """Export filtered cats data to XLSX file"""
        try:
            ui.notify('Export started...', type='info', position='top')
            
            # Get current filtered data
            filtered_cats = await apply_filters(reset_pagination=False)
            
            if not filtered_cats:
                ui.notify('No data to export', type='warning', position='top')
                return
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Cats Export"
            
            # Define headers (exclude actions column)
            headers = [col['label'] for col in cats_column if col['name'] != 'actions']
            
            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Write data rows
            for row_idx, cat in enumerate(filtered_cats, 2):
                # Format title
                title_display = cat.get('title')[0] if cat.get('title') and len(cat.get('title')) > 0 else ''
                
                # Format owner name
                owner_name = f"{cat.get('owner_firstname', '')} {cat.get('owner_surname', '')}".strip()
                
                # Format breeder name
                breeder_name = f"{cat.get('breed_firstname', '')} {cat.get('breed_surname', '')}".strip()
                
                # Format birthday
                birthday_display = cat.get('birthday').strftime('%Y-%m-%d') if cat.get('birthday') else ''
                
                # Create row data
                row_data = {
                    'id': cat.get('id'),
                    'firstname': cat.get('firstname', ''),
                    'surname': cat.get('surname', ''),
                    'callname': cat.get('callname', ''),
                    'gender': cat.get('gender', ''),
                    'birthday': birthday_display,
                    'microchip': cat.get('microchip', ''),
                    'title': title_display,
                    'eye_colour': cat.get('eye_colour', ''),
                    'hair_type': cat.get('hair_type', ''),
                    'colour': cat.get('colour', ''),
                    'breed_name': breeder_name,
                    'owner_name': owner_name,
                    'status': cat.get('status', ''),
                    'actions': ''
                }
                
                for col_idx, col_def in enumerate(cats_column, 1):
                    if col_def['name'] == 'actions':
                        continue
                    field_name = col_def['field']
                    value = row_data.get(field_name, '')
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Generate filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'cats_export_{timestamp}.xlsx'
            
            # Save to temporary file and read content
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                wb.save(tmp_file.name)
                temp_path = tmp_file.name
            
            # Read file content
            with open(temp_path, 'rb') as f:
                file_content = f.read()
            
            # Clean up temporary file
            os.unlink(temp_path)
            
            # Download the file
            ui.download(file_content, filename)
            
            ui.notify('XLSX file ready for download', type='positive', position='top')
            
        except Exception as e:
            print(f"Export error: {e}")
            ui.notify(f'Export failed: {str(e)}', type='negative', position='top')

    async def export_to_pdf():
        """Export filtered cats data to PDF file"""
        try:
            ui.notify('PDF export started...', type='info', position='top')
            
            # Get current filtered data
            filtered_cats = await apply_filters(reset_pagination=False)
            
            if not filtered_cats:
                ui.notify('No data to export', type='warning', position='top')
                return
            
            # Generate filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'cats_export_{timestamp}.pdf'
            
            # Create temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                temp_path = tmp_file.name
            
            # Create PDF document
            doc = SimpleDocTemplate(temp_path, pagesize=landscape(A4))
            story = []
            
            # Define styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            
            # Add title
            title = Paragraph("Cats Export", title_style)
            story.append(title)
            story.append(Spacer(1, 12))
            
            # Prepare table data (exclude actions column)
            headers = [col['label'] for col in cats_column if col['name'] != 'actions']
            table_data = [headers]
            
            # Add data rows
            for cat in filtered_cats:
                # Format title
                title_display = cat.get('title')[0] if cat.get('title') and len(cat.get('title')) > 0 else ''
                
                # Format owner name
                owner_name = f"{cat.get('owner_firstname', '')} {cat.get('owner_surname', '')}".strip()
                
                # Format breeder name
                breeder_name = f"{cat.get('breed_firstname', '')} {cat.get('breed_surname', '')}".strip()
                
                # Format birthday
                birthday_display = cat.get('birthday').strftime('%Y-%m-%d') if cat.get('birthday') else ''
                
                # Create row data
                row_data = {
                    'id': cat.get('id'),
                    'firstname': cat.get('firstname', ''),
                    'surname': cat.get('surname', ''),
                    'callname': cat.get('callname', ''),
                    'gender': cat.get('gender', ''),
                    'birthday': birthday_display,
                    'microchip': cat.get('microchip', ''),
                    'title': title_display,
                    'eye_colour': cat.get('eye_colour', ''),
                    'hair_type': cat.get('hair_type', ''),
                    'colour': cat.get('colour', ''),
                    'breed_name': breeder_name,
                    'owner_name': owner_name,
                    'status': cat.get('status', ''),
                    'actions': ''
                }
                
                pdf_row_data = []
                for col_def in cats_column:
                    if col_def['name'] == 'actions':
                        continue
                    field_name = col_def['field']
                    value = str(row_data.get(field_name, ''))
                    # Truncate long text for PDF display
                    if len(value) > 15:
                        value = value[:12] + "..."
                    pdf_row_data.append(value)
                table_data.append(pdf_row_data)
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                # Header row styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Data rows styling
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            story.append(table)
            
            # Build PDF
            doc.build(story)
            
            # Read file content
            with open(temp_path, 'rb') as f:
                file_content = f.read()
            
            # Clean up temporary file
            os.unlink(temp_path)
            
            # Download the file
            ui.download(file_content, filename)
            
            ui.notify('PDF file ready for download', type='positive', position='top')
            
        except Exception as e:
            print(f"PDF export error: {e}")
            ui.notify(f'PDF export failed: {str(e)}', type='negative', position='top')

    async def clear_all_filters():
        """Clear all filter inputs"""
        search_input.value = ''
        gender_filter.value = ''
        owner_filter_select.value = ''
        breeder_filter.value = ''
        eye_color_filter.value = ''
        hair_type_filter.value = ''
        status_filter.value = ''
        color_filter.value = ''
        birthday_from.value = ''
        birthday_to.value = ''
        weight_min.value = None
        weight_max.value = None
        breeding_animal_filter.value = ''
        breeding_lock_filter.value = ''
        await update_table()

    # Set up event handlers
    search_input.on_value_change(lambda e: update_table())
    gender_filter.on_value_change(lambda e: update_table())
    owner_filter_select.on_value_change(lambda e: update_table())
    breeder_filter.on_value_change(lambda e: update_table())
    eye_color_filter.on_value_change(lambda e: update_table())
    hair_type_filter.on_value_change(lambda e: update_table())
    status_filter.on_value_change(lambda e: update_table())
    color_filter.on_value_change(lambda e: update_table())
    birthday_from.on_value_change(lambda e: update_table())
    birthday_to.on_value_change(lambda e: update_table())
    weight_min.on_value_change(lambda e: update_table())
    weight_max.on_value_change(lambda e: update_table())
    breeding_animal_filter.on_value_change(lambda e: update_table())
    breeding_lock_filter.on_value_change(lambda e: update_table())
    clear_filters_btn.on_click(clear_all_filters)
    export_xlsx_btn.on_click(export_to_xlsx)
    export_pdf_btn.on_click(export_to_pdf)

    # Initial load
    await load_cats_data(offset=0, limit=PAGE_SIZE, reset=True)
    await update_table()
