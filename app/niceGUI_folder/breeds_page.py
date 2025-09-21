from datetime import datetime
import tempfile
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from app.niceGUI_folder.header import get_header
from app.niceGUI_folder.auth_middleware import require_auth
from nicegui import ui
from app.database_folder.orm import AsyncOrm


columns = [
    {'name': 'id', 'label': 'ID', 'field': 'id', 'align': 'left'},
    {'name': 'firstname', 'label': 'First Name', 'field': 'firstname', 'align': 'left'},
    {'name': 'surname', 'label': 'Surname', 'field': 'surname', 'align': 'left'},
    {'name': 'email', 'label': 'Email', 'field': 'email', 'align': 'left'},
    {'name': 'phone', 'label': 'Phone', 'field': 'phone', 'align': 'left'},
    {'name': 'gender', 'label': 'Gender', 'field': 'gender', 'align': 'left'},
    {'name': 'birthday', 'label': 'Birthday', 'field': 'birthday', 'align': 'left'},
    {'name': 'address', 'label': 'Address', 'field': 'address', 'align': 'left'},
    {'name': 'city', 'label': 'City', 'field': 'city', 'align': 'left'},
    {'name': 'country', 'label': 'Country', 'field': 'country', 'align': 'left'},
    {'name': 'zip', 'label': 'ZIP', 'field': 'zip', 'align': 'left'},
    {'name': 'description', 'label': 'Description', 'field': 'description', 'align': 'left'},
    {'name': 'actions', 'label': 'Actions', 'field': 'actions', 'align': 'center'},
]


def get_edit_button_vue():
    return r'''
      <q-tr :props="props">
        <q-td v-for="col in props.cols" :key="col.name" :props="props">
          <template v-if="col.name === 'actions'">
            <q-btn size="sm" color="primary" flat
                   :href="'/edit_breed/' + props.row.id"
                   label="Edit" />
          </template>
          <template v-else>
            {{ col.value }}
          </template>
        </q-td>
      </q-tr>
    '''


def breed_to_row(b):
    if isinstance(b, dict):
        return {
            'id': b.get('breed_id'),
            'firstname': b.get('breed_firstname'),
            'surname': b.get('breed_surname'),
            'email': b.get('breed_email'),
            'phone': b.get('breed_phone'),
            'gender': b.get('breed_gender'),
            'birthday': b.get('breed_birthday'),
            'address': b.get('breed_address'),
            'city': b.get('breed_city'),
            'country': b.get('breed_country'),
            'zip': b.get('breed_zip'),
            'description': b.get('breed_description'),
        }
    return {
        'id': getattr(b, 'breed_id', None),
        'firstname': getattr(b, 'breed_firstname', None),
        'surname': getattr(b, 'breed_surname', None),
        'email': getattr(b, 'breed_email', None),
        'phone': getattr(b, 'breed_phone', None),
        'gender': getattr(b, 'breed_gender', None),
        'birthday': getattr(b, 'breed_birthday', None),
        'address': getattr(b, 'breed_address', None),
        'city': getattr(b, 'breed_city', None),
        'country': getattr(b, 'breed_country', None),
        'zip': getattr(b, 'breed_zip', None),
        'description': getattr(b, 'breed_description', None),
    }


@ui.page('/breeds')
@require_auth(required_permission=1)
async def breeds_page_render(current_user=None, session_id=None):
    get_header('🐱 Breeds')

    # Pagination settings
    PAGE_SIZE = 100
    current_offset = 0
    all_breeds_data = []
    loading_more = False
    has_more_data = True

    async def load_breeds_data(offset=0, limit=PAGE_SIZE, reset=False):
        """Load breeds data with pagination"""
        nonlocal all_breeds_data, has_more_data, loading_more
        
        if loading_more:
            return []
            
        loading_more = True
        
        try:
            # Load data from database with pagination
            _, new_breeds = await AsyncOrm.get_breed(limit=limit, offset=offset)
            
            if reset:
                all_breeds_data = new_breeds
            else:
                all_breeds_data.extend(new_breeds)
            
            # Check if there's more data
            has_more_data = len(new_breeds) == limit
            
            return new_breeds
        finally:
            loading_more = False

    async def update_table():
        """Update the table with current data"""
        # Convert data to rows
        rows = [breed_to_row(b) for b in all_breeds_data]
        for row in rows:
            row['actions'] = ''

        # Update table
        table_container.clear()
        with table_container:
            if rows:
                table = ui.table(columns=columns, rows=rows, row_key='id').classes('q-pa-md')
                table.add_slot('body', get_edit_button_vue())
                
                # Add load more button if there's more data
                if has_more_data:
                    with ui.row().classes('q-pa-md justify-center'):
                        load_more_btn = ui.button('Load More', icon='expand_more',
                                                  on_click=load_more_data).props('color=primary outline')
                        if loading_more:
                            load_more_btn.props('loading')
                elif len(rows) > 0:
                    with ui.row().classes('q-pa-md justify-center'):
                        ui.label('All records loaded').classes('text-grey-6')
            else:
                ui.label('No breeds found').classes('text-center q-py-xl text-grey-6')

    async def load_more_data():
        """Load more data when button is clicked"""
        nonlocal current_offset, loading_more
        
        if loading_more or not has_more_data:
            return
            
        current_offset += PAGE_SIZE
        new_breeds = await load_breeds_data(offset=current_offset, limit=PAGE_SIZE, reset=False)
        
        if new_breeds:
            await update_table()

    with ui.row().classes('q-pa-md'):
        ui.button('Add Breed', on_click=lambda: ui.navigate.to('/add_breed')).classes('q-mr-sm')
        export_xlsx_btn = ui.button('Export XLSX', icon='download').props('color=primary').classes('q-mr-sm')
        export_pdf_btn = ui.button('Export PDF', icon='picture_as_pdf').props('color=secondary')

    # Table container
    table_container = ui.column()
    
    # Set up export event handlers
    export_xlsx_btn.on_click(lambda: export_to_xlsx(all_breeds_data))
    export_pdf_btn.on_click(lambda: export_to_pdf(all_breeds_data))

    # Initial load
    await load_breeds_data(offset=0, limit=PAGE_SIZE, reset=True)
    await update_table()


def export_to_xlsx(breeds_data):
    """Export breeds data to XLSX file"""
    try:
        ui.notify('Export started...', type='info', position='top')
        
        if not breeds_data:
            ui.notify('No data to export', type='warning', position='top')
            return
        
        # Convert data to rows
        rows = [breed_to_row(b) for b in breeds_data]
        for row in rows:
            row['actions'] = ''
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Breeds Export"
        
        # Define headers (exclude actions column)
        headers = [col['label'] for col in columns if col['name'] != 'actions']
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, col_def in enumerate(columns, 1):
                if col_def['name'] == 'actions':
                    continue
                field_name = col_def['field']
                value = row.get(field_name, '')
                if field_name == 'birthday' and value:
                    value = value.strftime('%Y-%m-%d') if hasattr(value, 'strftime') else str(value)
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'breeds_export_{timestamp}.xlsx'
        
        # Save to temporary file and read content
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            wb.save(tmp_file.name)
            temp_path = tmp_file.name
        
        # Read file content
        with open(temp_path, 'rb') as f:
            file_content = f.read()
        
        # Clean up temporary file
        os.unlink(temp_path)
        
        # Download the file
        ui.download(file_content, filename)
        
        ui.notify('XLSX file ready for download', type='positive', position='top')
        
    except Exception as e:
        print(f"Export error: {e}")
        ui.notify(f'Export failed: {str(e)}', type='negative', position='top')


def export_to_pdf(breeds_data):
    """Export breeds data to PDF file"""
    try:
        ui.notify('PDF export started...', type='info', position='top')
        
        if not breeds_data:
            ui.notify('No data to export', type='warning', position='top')
            return
        
        # Convert data to rows
        rows = [breed_to_row(b) for b in breeds_data]
        for row in rows:
            row['actions'] = ''
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'breeds_export_{timestamp}.pdf'
        
        # Create temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
            temp_path = tmp_file.name
        
        # Create PDF document
        doc = SimpleDocTemplate(temp_path, pagesize=landscape(A4))
        story = []
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Add title
        title = Paragraph("Breeds Export", title_style)
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Prepare table data (exclude actions column)
        headers = [col['label'] for col in columns if col['name'] != 'actions']
        table_data = [headers]
        
        # Add data rows
        for row in rows:
            row_data = []
            for col_def in columns:
                if col_def['name'] == 'actions':
                    continue
                field_name = col_def['field']
                value = str(row.get(field_name, ''))
                if field_name == 'birthday' and row.get(field_name):
                    birthday_val = row.get(field_name)
                    value = birthday_val.strftime('%Y-%m-%d') if hasattr(birthday_val, 'strftime') else str(birthday_val)
                # Truncate long text for PDF display
                if len(value) > 15:
                    value = value[:12] + "..."
                row_data.append(value)
            table_data.append(row_data)
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            # Header row styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        story.append(table)
        
        # Build PDF
        doc.build(story)
        
        # Read file content
        with open(temp_path, 'rb') as f:
            file_content = f.read()
        
        # Clean up temporary file
        os.unlink(temp_path)
        
        # Download the file
        ui.download(file_content, filename)
        
        ui.notify('PDF file ready for download', type='positive', position='top')
        
    except Exception as e:
        print(f"PDF export error: {e}")
        ui.notify(f'PDF export failed: {str(e)}', type='negative', position='top')
