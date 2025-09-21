from datetime import datetime
import tempfile
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from app.niceGUI_folder.header import get_header
from app.niceGUI_folder.auth_middleware import require_auth
from nicegui import ui
from app.database_folder.orm import AsyncOrm


columns = [
    {'name': 'id', 'label': 'ID', 'field': 'id', 'align': 'left'},
    {'name': 'firstname', 'label': 'First Name', 'field': 'firstname', 'align': 'left'},
    {'name': 'surname', 'label': 'Surname', 'field': 'surname', 'align': 'left'},
    {'name': 'email', 'label': 'Email', 'field': 'email', 'align': 'left'},
    {'name': 'phone', 'label': 'Phone', 'field': 'phone', 'align': 'left'},
    {'name': 'address', 'label': 'Address', 'field': 'address', 'align': 'left'},
    {'name': 'city', 'label': 'City', 'field': 'city', 'align': 'left'},
    {'name': 'country', 'label': 'Country', 'field': 'country', 'align': 'left'},
    {'name': 'zip', 'label': 'ZIP', 'field': 'zip', 'align': 'left'},
    {'name': 'birthday', 'label': 'Birthday', 'field': 'birthday', 'align': 'left'},
    {'name': 'permission', 'label': 'Permission', 'field': 'permission', 'align': 'left'},
    {'name': 'actions', 'label': 'Actions', 'field': 'actions', 'align': 'center'},
]


def get_edit_button_vue():
    return r'''
      <q-tr :props="props">
        <q-td v-for="col in props.cols" :key="col.name" :props="props">
          <template v-if="col.name === 'actions'">
            <q-btn size="sm" color="primary" flat
                   :href="'/edit_owner/' + props.row.id"
                   label="Edit" />
          </template>
          <template v-else>
            {{ col.value }}
          </template>
        </q-td>
      </q-tr>
    '''


def owner_to_row(o):
    if isinstance(o, dict):
        return {
            'id': o.get('owner_id'),
            'firstname': o.get('owner_firstname'),
            'surname': o.get('owner_surname'),
            'email': o.get('owner_email'),
            'phone': o.get('owner_phone'),
            'address': o.get('owner_address'),
            'city': o.get('owner_city'),
            'country': o.get('owner_country'),
            'zip': o.get('owner_zip'),
            'birthday': o.get('owner_birthday'),
            'permission': o.get('owner_permission'),
        }
    return {
        'id': getattr(o, 'owner_id', None),
        'firstname': getattr(o, 'owner_firstname', None),
        'surname': getattr(o, 'owner_surname', None),
        'email': getattr(o, 'owner_email', None),
        'phone': getattr(o, 'owner_phone', None),
        'address': getattr(o, 'owner_address', None),
        'city': getattr(o, 'owner_city', None),
        'country': getattr(o, 'owner_country', None),
        'zip': getattr(o, 'owner_zip', None),
        'birthday': getattr(o, 'owner_birthday', None),
        'permission': getattr(o, 'owner_permission', None),
    }


@ui.page('/owners')
@require_auth(required_permission=1)
async def owners_page_render(current_user=None, session_id=None):
    len_owners, owners = await AsyncOrm.get_owner()
    get_header('👤 Owners')

    with ui.row().classes('q-pa-md'):
        ui.button('Add Owner', on_click=lambda: ui.navigate.to('/add_owner')).classes('q-mr-sm')
        export_xlsx_btn = ui.button('Export XLSX', icon='download').props('color=primary').classes('q-mr-sm')
        export_pdf_btn = ui.button('Export PDF', icon='picture_as_pdf').props('color=secondary')

    rows = [owner_to_row(o) for o in (owners if isinstance(owners, list) else [owners])]

    for row in rows:
        row['actions'] = ''

    table = ui.table(columns=columns, rows=rows, row_key='id').classes('q-pa-md')
    table.add_slot('body', get_edit_button_vue())
    
    # Set up export event handlers
    export_xlsx_btn.on_click(lambda: export_to_xlsx(rows))
    export_pdf_btn.on_click(lambda: export_to_pdf(rows))


def export_to_xlsx(rows):
    """Export owners data to XLSX file"""
    try:
        ui.notify('Export started...', type='info', position='top')
        
        if not rows:
            ui.notify('No data to export', type='warning', position='top')
            return
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Owners Export"
        
        # Define headers (exclude actions column)
        headers = [col['label'] for col in columns if col['name'] != 'actions']
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, col_def in enumerate(columns, 1):
                if col_def['name'] == 'actions':
                    continue
                field_name = col_def['field']
                value = row.get(field_name, '')
                if field_name == 'birthday' and value:
                    value = value.strftime('%Y-%m-%d') if hasattr(value, 'strftime') else str(value)
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'owners_export_{timestamp}.xlsx'
        
        # Save to temporary file and read content
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            wb.save(tmp_file.name)
            temp_path = tmp_file.name
        
        # Read file content
        with open(temp_path, 'rb') as f:
            file_content = f.read()
        
        # Clean up temporary file
        os.unlink(temp_path)
        
        # Download the file
        ui.download(file_content, filename)
        
        ui.notify('XLSX file ready for download', type='positive', position='top')
        
    except Exception as e:
        print(f"Export error: {e}")
        ui.notify(f'Export failed: {str(e)}', type='negative', position='top')


def export_to_pdf(rows):
    """Export owners data to PDF file"""
    try:
        ui.notify('PDF export started...', type='info', position='top')
        
        if not rows:
            ui.notify('No data to export', type='warning', position='top')
            return
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'owners_export_{timestamp}.pdf'
        
        # Create temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
            temp_path = tmp_file.name
        
        # Create PDF document
        doc = SimpleDocTemplate(temp_path, pagesize=landscape(A4))
        story = []
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Add title
        title = Paragraph("Owners Export", title_style)
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Prepare table data (exclude actions column)
        headers = [col['label'] for col in columns if col['name'] != 'actions']
        table_data = [headers]
        
        # Add data rows
        for row in rows:
            row_data = []
            for col_def in columns:
                if col_def['name'] == 'actions':
                    continue
                field_name = col_def['field']
                value = str(row.get(field_name, ''))
                if field_name == 'birthday' and row.get(field_name):
                    birthday_val = row.get(field_name)
                    value = birthday_val.strftime('%Y-%m-%d') if hasattr(birthday_val, 'strftime') else str(birthday_val)
                # Truncate long text for PDF display
                if len(value) > 15:
                    value = value[:12] + "..."
                row_data.append(value)
            table_data.append(row_data)
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            # Header row styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        story.append(table)
        
        # Build PDF
        doc.build(story)
        
        # Read file content
        with open(temp_path, 'rb') as f:
            file_content = f.read()
        
        # Clean up temporary file
        os.unlink(temp_path)
        
        # Download the file
        ui.download(file_content, filename)
        
        ui.notify('PDF file ready for download', type='positive', position='top')
        
    except Exception as e:
        print(f"PDF export error: {e}")
        ui.notify(f'PDF export failed: {str(e)}', type='negative', position='top')
